"""
MF Fund Ranker — DUAL ENGINE Model V2
======================================
Improvements:
  ✓ 2Y CAGR column added in output (between 1Y and 3Y CAGR)
  ✓ 4-row header: Title | Info | Group banners | Column headers
  ✓ Group banners: ◄ Momentum ► | ◄ Long-Term ► | ◄ Engine Scores ►
  ✓ Row 2 dynamically shows model config pulled from assumption constants
  ✓ Assumptions sheet with improved color scheme
"""

import pandas as pd
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ───────────────────────────────────────────────────────────────────
INPUT_FILE  = "dashboard_data.xlsx"
OUTPUT_FILE = "mf_ranked_screener.xlsx"

QUALITY_FILTERS = {
    "cagr_2y_min": 0.10,
    "cagr_3y_min": 0.12,
}

ENGINE1_WEIGHTS = {
    "return_1y":  0.25,
    "return_2y":  0.30,
    "return_3y":  0.45,
}

ENGINE2_WEIGHTS = {
    "return_6m":  0.30,
    "return_3m":  0.20,
    "return_1y":  0.25,
    "return_1m":  0.25,
}

COMPOSITE_BLEND = {
    "engine1_quality":  0.45,
    "engine2_momentum": 0.55,
}

TREND_BONUS = 5.0

FILTERS = {
    "cat_level_1": "Open Ended Schemes",
    "cat_level_2": "Equity Scheme",
    "plan_type":   "Regular",
    "option_type": "Growth",
}

COLUMN_MAP = {
    "scheme_name": "scheme_name",
    "category":    "cat_level_3",
    "amc":         "amc_name",
    "return_1m":   "return_30d",
    "return_3m":   "return_90d",
    "return_6m":   "return_180d",
    "return_1y":   "return_365d",
    "return_2y":   "return_730d",
    "return_3y":   "return_1095d",
    "nav":         "latest_nav",
}

# ── COLORS ───────────────────────────────────────────────────────────────────
C = {
    # Structural
    "title_bg":      "0D1117",   # near-black
    "title_fg":      "FFFFFF",
    "info_bg":       "F0F4F8",
    "info_fg":       "445566",
    "border":        "D0D8E4",
    # Header rows
    "col_hdr_bg":    "1A3A5C",   # navy
    "col_hdr_fg":    "FFFFFF",
    # Group banner colours
    "momentum_bg":   "E8560A",   # orange  (Momentum group)
    "momentum_fg":   "FFFFFF",
    "longterm_bg":   "1E6B8C",   # teal    (Long-Term group)
    "longterm_fg":   "FFFFFF",
    "engine_bg":     "2D5016",   # forest  (Engine Scores group)
    "engine_fg":     "FFFFFF",
    # Data row backgrounds
    "rank1_bg":      "FFD700",
    "rank2_bg":      "E8E8E8",
    "rank3_bg":      "D4956A",
    "alt_row":       "F5F8FC",
    "white":         "FFFFFF",
    # Return colours
    "positive":      "1E7A4B",
    "negative":      "C0392B",
    # Engine col tints
    "engine1_tint":  "E3F2FD",   # light blue
    "engine2_tint":  "FFF3E0",   # light amber
    "comp_tint":     "F0FFF4",   # light green
    # Assumption sheet palette
    "asmp_title":    "0D1117",
    "asmp_e1":       "1E3A5F",
    "asmp_e2":       "7B3F00",
    "asmp_blend":    "2D5016",
    "asmp_signal":   "4A235A",
    "asmp_row_e1":   "EBF5FB",
    "asmp_row_e2":   "FEF9E7",
    "asmp_row_bl":   "EAFAF1",
    "asmp_row_sg":   "F5EEF8",
}

# ── DATA LOADING ─────────────────────────────────────────────────────────────
def apply_filters(df):
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for col_key, value in FILTERS.items():
        actual = cols_lower.get(col_key.lower().strip())
        if actual:
            df = df[df[actual].astype(str).str.strip().str.lower() == str(value).strip().lower()]
    return df

def load_data():
    sheets = pd.read_excel(INPUT_FILE, sheet_name=None)
    frames = [df for df in sheets.values() if "scheme_name" in df.columns]
    combined = pd.concat(frames, ignore_index=True)
    return apply_filters(combined)

# ── SCORING ───────────────────────────────────────────────────────────────────
def to_num(series):
    s = series.astype(str).str.replace('%', '').str.replace(',', '').str.strip()
    return pd.to_numeric(s, errors='coerce')

def pct_rank(series):
    ranks = series.rank(method='min', na_option='bottom')
    return (ranks - 1) / max(len(series) - 1, 1) * 100

def cagr_2y(r):
    if pd.isna(r): return float('nan')
    return ((1 + float(r) / 100) ** 0.5 - 1) * 100

def score_funds(df):
    df = df.copy()
    df["_r1m"]     = to_num(df[COLUMN_MAP["return_1m"]])
    df["_r3m"]     = to_num(df[COLUMN_MAP["return_3m"]])
    df["_r6m"]     = to_num(df[COLUMN_MAP["return_6m"]])
    df["_r1y"]     = to_num(df[COLUMN_MAP["return_1y"]])
    df["_r2y_raw"] = to_num(df[COLUMN_MAP["return_2y"]])
    df["_r2y_cagr"]= df["_r2y_raw"].apply(cagr_2y)
    df["_r3y"]     = to_num(df[COLUMN_MAP["return_3y"]])
    df["_cat"]     = df[COLUMN_MAP["category"]].astype(str).str.strip().str.title()

    # Quality gate
    mask_2y = df["_r2y_cagr"] > (QUALITY_FILTERS["cagr_2y_min"] * 100)
    mask_3y = df["_r3y"]      > (QUALITY_FILTERS["cagr_3y_min"] * 100)
    quality  = mask_2y & mask_3y
    df["_qualifies"] = quality

    # Trend signal
    trend = (df["_r6m"] > df["_r3m"]) & (df["_r3m"] > df["_r1m"])
    df["_trend"] = trend.map({True: "📈 Uptrend", False: ""})

    df["_e1"] = 0.0
    df["_e2"] = 0.0

    for cat in df["_cat"].unique():
        cm  = df["_cat"] == cat
        cq  = cm & quality

        # Engine 1 — quality (only qualified funds)
        if cq.sum() > 0:
            e1 = (pct_rank(df.loc[cq, "_r1y"])     * ENGINE1_WEIGHTS["return_1y"] +
                  pct_rank(df.loc[cq, "_r2y_cagr"]) * ENGINE1_WEIGHTS["return_2y"] +
                  pct_rank(df.loc[cq, "_r3y"])      * ENGINE1_WEIGHTS["return_3y"])
            df.loc[cq, "_e1"] = e1

        # Engine 2 — momentum (all funds in category)
        e2 = (pct_rank(df.loc[cm, "_r6m"]) * ENGINE2_WEIGHTS["return_6m"] +
              pct_rank(df.loc[cm, "_r3m"]) * ENGINE2_WEIGHTS["return_3m"] +
              pct_rank(df.loc[cm, "_r1y"]) * ENGINE2_WEIGHTS["return_1y"] +
              pct_rank(df.loc[cm, "_r1m"]) * ENGINE2_WEIGHTS["return_1m"])
        up_idx = df.loc[cm][df.loc[cm, "_trend"] == "📈 Uptrend"].index
        e2.loc[e2.index.isin(up_idx)] += TREND_BONUS
        df.loc[cm, "_e2"] = e2.clip(0, 100)

    df["_comp"] = (df["_e1"] * COMPOSITE_BLEND["engine1_quality"] +
                   df["_e2"] * COMPOSITE_BLEND["engine2_momentum"])

    df["_rank"] = (df.groupby("_cat")["_comp"]
                     .rank(method='min', ascending=False)
                     .astype(int))
    return df

# ── FORMATTING HELPERS ────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def border():
    s = Side(style='thin', color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_pct(val):
    if pd.isna(val) or val is None: return "—"
    try: return f"{float(val):+.2f}%"
    except: return "—"

def score_col(val):
    try: v = float(val)
    except: return "888888"
    if v >= 75: return "1E7A4B"
    if v >= 50: return "E67E00"
    return "C0392B"

def signal(comp):
    if comp >= 75: return "⭐ Strong Buy"
    if comp >= 55: return "✅ Buy"
    return "⚠️ Watch"

def clean_name(name):
    return re.sub(r'[\\/*?:\[\]]', '', str(name))[:31]

def hfont(size=10, color="FFFFFF", bold=True):
    return Font(name="Arial", bold=bold, size=size, color=color)

def dfont(size=9, bold=False, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def set_cell(ws, row, col, value, font=None, fill_=None, align=None, bdr=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:  c.font   = font
    if fill_: c.fill   = fill_
    if align: c.alignment = align
    if bdr:   c.border = bdr
    return c

CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
VCENTER  = Alignment(horizontal="center", vertical="center")

# ── DYNAMIC ROW 2 TEXT ────────────────────────────────────────────────────────
def row2_text(cat, total_funds, qualified_funds):
    e1_pct = int(COMPOSITE_BLEND["engine1_quality"]  * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_momentum"] * 100)
    f2y    = int(QUALITY_FILTERS["cagr_2y_min"]      * 100)
    f3y    = int(QUALITY_FILTERS["cagr_3y_min"]      * 100)
    return (
        f"Dual Engine Model: {e1_pct}% Quality (LT) + {e2_pct}% Momentum (ST)  |  "
        f"Quality Gate: 2Y CAGR >{f2y}% & 3Y CAGR >{f3y}%  |  "
        f"Total funds: {total_funds}  |  Passed quality gate: {qualified_funds}"
    )

# ── COLUMN LAYOUT ─────────────────────────────────────────────────────────────
# Col:  1=Rank  2=Name  3=AMC  4=1M  5=3M  6=6M  7=1Y  8=2Y CAGR  9=3Y CAGR  10=E1  11=E2  12=Comp
MOMENTUM_COLS  = (4, 6)   # 1M → 6M
LONGTERM_COLS  = (7, 9)   # 1Y → 3Y CAGR
ENGINE_COLS    = (10, 12) # E1 → Composite

COL_HEADERS = [
    "Rank", "Scheme Name", "AMC",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Quality)", "Engine 2\n(Momentum)", "Composite\nScore",
]

COL_WIDTHS = [6, 50, 20, 9, 9, 9, 9, 9, 9, 14, 14, 12]

RETURN_COLS_IDX = {4, 5, 6, 7, 8, 9}   # columns with % returns
ENGINE_COLS_IDX = {10, 11, 12}


# ── BUILD A CATEGORY SHEET ───────────────────────────────────────────────────
def build_category_sheet(wb, cat, cat_df):
    ws = wb.create_sheet(clean_name(cat))
    bd = border()

    total     = len(cat_df)
    qualified = int(cat_df["_qualifies"].sum())

    # ── ROW 1: Title ──────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COL_HEADERS))}1")
    c = ws["A1"]
    c.value     = f"✦  {cat.upper()}  ✦"
    c.font      = Font(name="Arial", bold=True, size=14, color=C["title_fg"])
    c.fill      = fill(C["title_bg"])
    c.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # ── ROW 2: Dynamic info ───────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(COL_HEADERS))}2")
    c = ws["A2"]
    c.value     = row2_text(cat, total, qualified)
    c.font      = Font(name="Arial", italic=True, size=8.5, color=C["info_fg"])
    c.fill      = fill(C["info_bg"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # ── ROW 3: Group banners ──────────────────────────────────────────────
    # Cols 1-3: blank (Rank, Name, AMC)
    for ci in range(1, 4):
        c = ws.cell(row=3, column=ci)
        c.fill = fill(C["col_hdr_bg"])
        c.border = bd

    # Momentum group
    ws.merge_cells(f"{get_column_letter(MOMENTUM_COLS[0])}3:{get_column_letter(MOMENTUM_COLS[1])}3")
    c = ws.cell(row=3, column=MOMENTUM_COLS[0])
    c.value     = "◄  Momentum  ►"
    c.font      = hfont(size=9, color=C["momentum_fg"])
    c.fill      = fill(C["momentum_bg"])
    c.alignment = CENTER
    c.border    = bd
    for ci in range(MOMENTUM_COLS[0]+1, MOMENTUM_COLS[1]+1):
        ws.cell(row=3, column=ci).fill   = fill(C["momentum_bg"])
        ws.cell(row=3, column=ci).border = bd

    # Long-Term group
    ws.merge_cells(f"{get_column_letter(LONGTERM_COLS[0])}3:{get_column_letter(LONGTERM_COLS[1])}3")
    c = ws.cell(row=3, column=LONGTERM_COLS[0])
    c.value     = "◄  Long-Term  ►"
    c.font      = hfont(size=9, color=C["longterm_fg"])
    c.fill      = fill(C["longterm_bg"])
    c.alignment = CENTER
    c.border    = bd
    for ci in range(LONGTERM_COLS[0]+1, LONGTERM_COLS[1]+1):
        ws.cell(row=3, column=ci).fill   = fill(C["longterm_bg"])
        ws.cell(row=3, column=ci).border = bd

    # Engine Scores group
    ws.merge_cells(f"{get_column_letter(ENGINE_COLS[0])}3:{get_column_letter(ENGINE_COLS[1])}3")
    c = ws.cell(row=3, column=ENGINE_COLS[0])
    c.value     = "◄  Engine Scores  ►"
    c.font      = hfont(size=9, color=C["engine_fg"])
    c.fill      = fill(C["engine_bg"])
    c.alignment = CENTER
    c.border    = bd
    for ci in range(ENGINE_COLS[0]+1, ENGINE_COLS[1]+1):
        ws.cell(row=3, column=ci).fill   = fill(C["engine_bg"])
        ws.cell(row=3, column=ci).border = bd

    ws.row_dimensions[3].height = 18

    # ── ROW 4: Column headers ─────────────────────────────────────────────
    for ci, hdr in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font      = hfont(size=9)
        c.fill      = fill(C["col_hdr_bg"])
        c.alignment = CENTER
        c.border    = bd
    ws.row_dimensions[4].height = 28

    # ── DATA ROWS (start row 5) ───────────────────────────────────────────
    for i, (_, row) in enumerate(cat_df.iterrows(), 5):
        rank = row["_rank"]
        if   rank == 1: rbg = C["rank1_bg"]
        elif rank == 2: rbg = C["rank2_bg"]
        elif rank == 3: rbg = C["rank3_bg"]
        elif i % 2 == 0: rbg = C["alt_row"]
        else:            rbg = C["white"]

        e1   = round(row["_e1"],   1)
        e2   = round(row["_e2"],   1)
        comp = round(row["_comp"], 1)

        vals = [
            rank,
            row.get(COLUMN_MAP["scheme_name"], "—"),
            row.get(COLUMN_MAP["amc"], "—"),
            fmt_pct(row["_r1m"]),
            fmt_pct(row["_r3m"]),
            fmt_pct(row["_r6m"]),
            fmt_pct(row["_r1y"]),
            fmt_pct(row["_r2y_cagr"]),
            fmt_pct(row["_r3y"]),
            e1, e2, comp,
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border    = bd
            c.alignment = LEFT if ci == 2 else CENTER

            if ci in RETURN_COLS_IDX and isinstance(val, str) and val != "—":
                num = float(val.replace('%', ''))
                c.font = Font(name="Arial", bold=(rank <= 3), size=9,
                              color=C["positive"] if num >= 0 else C["negative"])
                c.fill = fill(rbg)
            elif ci == 10:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C["engine1_tint"])
            elif ci == 11:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C["engine2_tint"])
            elif ci == 12:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C["comp_tint"])
            else:
                c.font = dfont(bold=(rank <= 3))
                c.fill = fill(rbg)

        ws.row_dimensions[i].height = 16

    # Column widths
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"
    return ws


# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────
def build_summary(wb, df_scored, categories):
    ws = wb.create_sheet("🏆 SUMMARY", 0)
    bd = border()
    ncols = len(COL_HEADERS)

    # Row 1: Master title
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = "MF INTELLIGENCE  —  DUAL ENGINE RANKED SCREENER"
    c.font      = Font(name="Arial", bold=True, size=15, color=C["title_fg"])
    c.fill      = fill(C["title_bg"])
    c.alignment = CENTER
    ws.row_dimensions[1].height = 34

    # Row 2: Global summary info (dynamic)
    e1_pct = int(COMPOSITE_BLEND["engine1_quality"]  * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_momentum"] * 100)
    f2y    = int(QUALITY_FILTERS["cagr_2y_min"]      * 100)
    f3y    = int(QUALITY_FILTERS["cagr_3y_min"]      * 100)
    total_funds = len(df_scored)
    qual_funds  = int(df_scored["_qualifies"].sum())
    info_text = (
        f"Composite: {e1_pct}% Engine 1 (Quality: 2Y CAGR>{f2y}%, 3Y CAGR>{f3y}%)  +  "
        f"{e2_pct}% Engine 2 (Momentum: 6M/3M/1M trend)  |  "
        f"Total: {total_funds} funds  |  Quality gate passed: {qual_funds}  |  "
        f"Categories: {len(categories)}"
    )
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value     = info_text
    c.font      = Font(name="Arial", italic=True, size=8.5, color=C["info_fg"])
    c.fill      = fill(C["info_bg"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Row 3: Group banners (same as category sheets)
    for ci in range(1, 4):
        ws.cell(row=3, column=ci).fill   = fill(C["col_hdr_bg"])
        ws.cell(row=3, column=ci).border = bd

    ws.merge_cells(f"{get_column_letter(MOMENTUM_COLS[0])}3:{get_column_letter(MOMENTUM_COLS[1])}3")
    c = ws.cell(row=3, column=MOMENTUM_COLS[0], value="◄  Momentum  ►")
    c.font = hfont(size=9, color=C["momentum_fg"]); c.fill = fill(C["momentum_bg"]); c.alignment = CENTER; c.border = bd
    for ci in range(MOMENTUM_COLS[0]+1, MOMENTUM_COLS[1]+1):
        ws.cell(row=3, column=ci).fill = fill(C["momentum_bg"]); ws.cell(row=3, column=ci).border = bd

    ws.merge_cells(f"{get_column_letter(LONGTERM_COLS[0])}3:{get_column_letter(LONGTERM_COLS[1])}3")
    c = ws.cell(row=3, column=LONGTERM_COLS[0], value="◄  Long-Term  ►")
    c.font = hfont(size=9, color=C["longterm_fg"]); c.fill = fill(C["longterm_bg"]); c.alignment = CENTER; c.border = bd
    for ci in range(LONGTERM_COLS[0]+1, LONGTERM_COLS[1]+1):
        ws.cell(row=3, column=ci).fill = fill(C["longterm_bg"]); ws.cell(row=3, column=ci).border = bd

    ws.merge_cells(f"{get_column_letter(ENGINE_COLS[0])}3:{get_column_letter(ENGINE_COLS[1])}3")
    c = ws.cell(row=3, column=ENGINE_COLS[0], value="◄  Engine Scores  ►")
    c.font = hfont(size=9, color=C["engine_fg"]); c.fill = fill(C["engine_bg"]); c.alignment = CENTER; c.border = bd
    for ci in range(ENGINE_COLS[0]+1, ENGINE_COLS[1]+1):
        ws.cell(row=3, column=ci).fill = fill(C["engine_bg"]); ws.cell(row=3, column=ci).border = bd

    ws.row_dimensions[3].height = 18

    # Row 4: Column headers
    for ci, hdr in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont(size=9); c.fill = fill(C["col_hdr_bg"]); c.alignment = CENTER; c.border = bd
    ws.row_dimensions[4].height = 28

    # Data rows: one per category (top-ranked fund)
    for i, cat in enumerate(categories, 5):
        cat_df = df_scored[df_scored["_cat"] == cat].sort_values("_rank")
        if cat_df.empty: continue
        top  = cat_df.iloc[0]
        comp = top["_comp"]
        rbg  = C["alt_row"] if i % 2 == 0 else C["white"]

        vals = [
            "1",
            top.get(COLUMN_MAP["scheme_name"], "—"),
            top.get(COLUMN_MAP["amc"], "—"),
            fmt_pct(top["_r1m"]),
            fmt_pct(top["_r3m"]),
            fmt_pct(top["_r6m"]),
            fmt_pct(top["_r1y"]),
            fmt_pct(top["_r2y_cagr"]),
            fmt_pct(top["_r3y"]),
            round(top["_e1"], 1),
            round(top["_e2"], 1),
            round(comp, 1),
        ]

        # Insert category name in rank column as label
        vals[0] = cat

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border    = bd
            c.alignment = LEFT if ci in {1, 2} else CENTER

            if ci in RETURN_COLS_IDX and isinstance(val, str) and val != "—":
                num = float(val.replace('%', ''))
                c.font = Font(name="Arial", size=9, color=C["positive"] if num >= 0 else C["negative"])
                c.fill = fill(rbg)
            elif ci == 10:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val)); c.fill = fill(C["engine1_tint"])
            elif ci == 11:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val)); c.fill = fill(C["engine2_tint"])
            elif ci == 12:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val)); c.fill = fill(C["comp_tint"])
            else:
                c.font = dfont(); c.fill = fill(rbg)

        # Signal in last cell (col 12 overlay — add as annotation in next col if needed)
        ws.row_dimensions[i].height = 18

    # Signal column header + data
    sig_col = ncols + 1
    ws.cell(row=4, column=sig_col, value="Signal").font   = hfont(size=9)
    ws.cell(row=4, column=sig_col).fill                   = fill(C["col_hdr_bg"])
    ws.cell(row=4, column=sig_col).alignment              = CENTER
    ws.cell(row=4, column=sig_col).border                 = bd
    ws.column_dimensions[get_column_letter(sig_col)].width = 14

    for i, cat in enumerate(categories, 5):
        cat_df = df_scored[df_scored["_cat"] == cat].sort_values("_rank")
        if cat_df.empty: continue
        top  = cat_df.iloc[0]
        comp = top["_comp"]
        rbg  = C["alt_row"] if i % 2 == 0 else C["white"]
        sig  = signal(comp)
        c = ws.cell(row=i, column=sig_col, value=sig)
        c.font      = Font(name="Arial", bold=True, size=9)
        c.fill      = fill(rbg)
        c.border    = bd
        c.alignment = CENTER

    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"


# ── ASSUMPTIONS SHEET ─────────────────────────────────────────────────────────
def build_assumptions(wb, df_scored):
    ws = wb.create_sheet("📋 ASSUMPTIONS", 1)
    bd = border()

    def section(row, title, bg, fg="FFFFFF", ncols=3):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws[f"A{row}"]
        c.value     = title
        c.font      = Font(name="Arial", bold=True, size=11, color=fg)
        c.fill      = fill(bg)
        c.alignment = CENTER
        ws.row_dimensions[row].height = 22
        return row + 1

    def kv_row(row, key, val, desc="", row_bg="FFFFFF"):
        cells = [(1, key), (2, val), (3, desc)]
        for ci, text in cells:
            c = ws.cell(row=row, column=ci, value=text)
            c.font      = Font(name="Arial", size=9)
            c.fill      = fill(row_bg)
            c.border    = bd
            c.alignment = LEFT if ci in {1, 3} else CENTER
        ws.row_dimensions[row].height = 15
        return row + 1

    def col_headers(row, headers, bg):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c.fill      = fill(bg)
            c.border    = bd
            c.alignment = CENTER
        ws.row_dimensions[row].height = 16
        return row + 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"] = "DUAL ENGINE MODEL — ASSUMPTIONS & METHODOLOGY"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = fill(C["asmp_title"])
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Source: {INPUT_FILE}"
    ws["A2"].font      = Font(name="Arial", italic=True, size=8.5, color=C["info_fg"])
    ws["A2"].fill      = fill(C["info_bg"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 14

    r = 4

    # ── COMPOSITE BLEND ────────────────────────────────────────────────────
    r = section(r, "⚖  COMPOSITE BLEND", C["asmp_blend"])
    r = col_headers(r, ["Parameter", "Value", "Description"], C["asmp_blend"])
    r = kv_row(r, "Engine 1 Weight (Quality)",  f"{int(COMPOSITE_BLEND['engine1_quality']*100)}%",
               "Long-term structural soundness", C["asmp_row_bl"])
    r = kv_row(r, "Engine 2 Weight (Momentum)", f"{int(COMPOSITE_BLEND['engine2_momentum']*100)}%",
               "Short-term trend strength",      C["asmp_row_bl"])
    r += 1

    # ── ENGINE 1 ───────────────────────────────────────────────────────────
    r = section(r, "🔵  ENGINE 1: LONG-TERM QUALITY", C["asmp_e1"])
    r = col_headers(r, ["Metric", "Weight", "Notes"], C["asmp_e1"])
    for metric, w in ENGINE1_WEIGHTS.items():
        label = {"return_1y": "1Y Return", "return_2y": "2Y CAGR (annualised from 2Y total)", "return_3y": "3Y CAGR"}[metric]
        note  = {"return_1y": "Recent 1-year performance", "return_2y": "√(1+r)−1 applied to 2Y total return", "return_3y": "Primary long-term signal (highest weight)"}[metric]
        r = kv_row(r, label, f"{int(w*100)}%", note, C["asmp_row_e1"])
    r += 1

    r = section(r, "🔵  ENGINE 1: QUALITY GATES (Hard Filters)", C["asmp_e1"])
    r = col_headers(r, ["Filter", "Threshold", "Effect"], C["asmp_e1"])
    r = kv_row(r, "2Y CAGR Minimum", f">{int(QUALITY_FILTERS['cagr_2y_min']*100)}%",
               "Funds below receive 0 Engine 1 score", C["asmp_row_e1"])
    r = kv_row(r, "3Y CAGR Minimum", f">{int(QUALITY_FILTERS['cagr_3y_min']*100)}%",
               "Funds below receive 0 Engine 1 score", C["asmp_row_e1"])

    total = len(df_scored)
    qual  = int(df_scored["_qualifies"].sum())
    r = kv_row(r, "Current run — funds qualified", f"{qual} / {total}",
               f"{qual/total*100:.1f}% of filtered universe passed", C["asmp_row_e1"])
    r += 1

    # ── ENGINE 2 ───────────────────────────────────────────────────────────
    r = section(r, "🟠  ENGINE 2: SHORT-TERM MOMENTUM", C["asmp_e2"])
    r = col_headers(r, ["Metric", "Weight", "Notes"], C["asmp_e2"])
    for metric, w in ENGINE2_WEIGHTS.items():
        label = {"return_6m": "6M Return", "return_3m": "3M Return", "return_1y": "1Y Return", "return_1m": "1M Return"}[metric]
        note  = {"return_6m": "Primary momentum signal", "return_3m": "Mid-range trend", "return_1y": "Bridge metric", "return_1m": "Latest directional signal"}[metric]
        r = kv_row(r, label, f"{int(w*100)}%", note, C["asmp_row_e2"])
    r += 1

    r = section(r, "🟠  TREND CONFIRMATION BONUS", C["asmp_e2"])
    r = col_headers(r, ["Condition", "Bonus", "Rationale"], C["asmp_e2"])
    r = kv_row(r, "6M > 3M > 1M Return", f"+{int(TREND_BONUS)} pts",
               "Confirms progressive momentum (capped at 100)", C["asmp_row_e2"])

    trend_count = int((df_scored["_trend"] == "📈 Uptrend").sum())
    r = kv_row(r, "Funds with uptrend this run", str(trend_count),
               f"{trend_count/total*100:.1f}% of filtered universe", C["asmp_row_e2"])
    r += 1

    # ── SCORING LOGIC ─────────────────────────────────────────────────────
    r = section(r, "📐  SCORING & RANKING LOGIC", C["col_hdr_bg"])
    r = col_headers(r, ["Step", "Detail", "Scope"], C["col_hdr_bg"])
    steps = [
        ("1. Percentile rank", "All return metrics ranked 0–100 within category", "Per category"),
        ("2. Weighted sum",    "Engine scores = Σ (percentile × weight)",         "Per engine"),
        ("3. Trend bonus",     f"+{int(TREND_BONUS)} added to Engine 2 if 6M>3M>1M", "Engine 2 only"),
        ("4. Composite",       f"{int(COMPOSITE_BLEND['engine1_quality']*100)}% E1 + {int(COMPOSITE_BLEND['engine2_momentum']*100)}% E2", "Final score"),
        ("5. Category rank",   "Funds sorted desc by Composite within each category", "Per category"),
    ]
    for s in steps:
        r = kv_row(r, s[0], s[1], s[2], C["white"] if steps.index(s) % 2 == 0 else C["alt_row"])
    r += 1

    # ── SIGNALS ───────────────────────────────────────────────────────────
    r = section(r, "🚦  SIGNAL INTERPRETATION", C["asmp_signal"])
    r = col_headers(r, ["Signal", "Composite Threshold", "Meaning"], C["asmp_signal"])
    sigs = [
        ("⭐ Strong Buy", "≥ 75", "Top-tier quality + momentum"),
        ("✅ Buy",        "≥ 55", "Solid score across both engines"),
        ("⚠️ Watch",      "< 55", "Below target or failed quality gate"),
    ]
    for sg in sigs:
        r = kv_row(r, sg[0], sg[1], sg[2], C["asmp_row_sg"])
    r += 1

    # ── DATA FILTERS ──────────────────────────────────────────────────────
    r = section(r, "🔎  DATA FILTERS APPLIED", C["asmp_blend"])
    r = col_headers(r, ["Filter Field", "Value", "Effect"], C["asmp_blend"])
    for k, v in FILTERS.items():
        r = kv_row(r, k, str(v), "Rows not matching are excluded", C["asmp_row_bl"])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 44


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading data...")
    df = load_data()
    print(f"  {len(df)} rows after filters")

    print("Scoring funds...")
    df_scored = score_funds(df)

    categories = sorted(df_scored["_cat"].unique())
    print(f"  {len(categories)} categories: {categories}")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    print("Building Summary sheet...")
    build_summary(wb, df_scored, categories)

    print("Building Assumptions sheet...")
    build_assumptions(wb, df_scored)

    print("Building category sheets...")
    for cat in categories:
        cat_df = df_scored[df_scored["_cat"] == cat].sort_values("_rank")
        if cat_df.empty: continue
        build_category_sheet(wb, cat, cat_df)
        print(f"  ✓ {cat} ({len(cat_df)} funds)")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved → {OUTPUT_FILE}")

    # Quick top picks
    print("\nTop fund per category:")
    for cat in categories:
        top = df_scored[df_scored["_cat"] == cat].sort_values("_rank").iloc[0]
        print(f"  {cat:30s}  Composite={top['_comp']:.1f}  "
              f"E1={top['_e1']:.0f}  E2={top['_e2']:.0f}  "
              f"{signal(top['_comp'])}  |  {top.get(COLUMN_MAP['scheme_name'],'')[:55]}")


if __name__ == "__main__":
    main()
