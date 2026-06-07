import pandas as pd
import re
from datetime import datetime
from typing import Tuple
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
@dataclass
class Config:
    INPUT_FILE: str = "dashboard_data.xlsx"
    OUTPUT_FILE: str = "mf_ranked_screener.xlsx"
    TREND_BONUS: float = 5.0

CONFIG = Config()

QUALITY_FILTERS = {
    "cagr_2y_min": 0.10,
    "cagr_3y_min": 0.12,
    "min_1y_return": -30.0,
}

ENGINE1_WEIGHTS = {
    "return_6m": 0.30,
    "return_3m": 0.20,
    "return_1y": 0.25,
    "return_1m": 0.25,
}

ENGINE2_WEIGHTS = {
    "return_1y": 0.25,
    "return_2y": 0.30,
    "return_3y": 0.45,
}

COMPOSITE_BLEND = {
    "engine1_momentum": 0.55,
    "engine2_quality": 0.45,
}

FILTERS = {
    "cat_level_1": "Open Ended Schemes",
    "cat_level_2": "Other Scheme",
    "plan_type": "Regular",
    "option_type": "Growth",
}

COLUMN_MAP = {
    "scheme_name": "scheme_name",
    "category": "cat_level_3",
    "amc": "amc_name",
    "return_1m": "return_30d",
    "return_3m": "return_90d",
    "return_6m": "return_180d",
    "return_1y": "return_365d",
    "return_2y": "return_730d",
    "return_3y": "return_1095d",
    "nav": "latest_nav",
}

# ══════════════════════════════════════════════════════════════════════════════
# ★ FIX #1 — TRUE PERCENT NUMBER FORMAT
# Excel '%' multiplies by 100. Returns are now stored as DECIMAL FRACTIONS
# (e.g. 28.04% -> 0.2804) and displayed with a genuine percent format so the
# stored value and the displayed value are CONSISTENT, sortable and chartable.
# ══════════════════════════════════════════════════════════════════════════════
PCT_FMT = '+0.00%;-0.00%;0.00%'   # ★ FIX: real percent format (value stored as fraction)
SCORE_FMT = '0.0'

# ══════════════════════════════════════════════════════════════════════════════
# ASSET TAGGING RULES - Priority-Ordered
# ══════════════════════════════════════════════════════════════════════════════
ASSET_TAG_RULES = [
    {"tag": "Silver", "contains": ["silver"], "priority": 1},
    {"tag": "Gold", "contains": ["gold"], "priority": 2},
    {"tag": "G-SEC", "contains": ["g-sec", "gsec"], "priority": 3},
    {"tag": "GILT", "contains": ["gilt"], "priority": 4},
    {"tag": "NASDAQ", "contains": ["nasdaq", "nq100", "nq 100"], "priority": 5},
    {"tag": "S&P 500", "contains": ["s&p 500", "s&p500", "s&p 50"], "priority": 6},
    {"tag": "International", "contains": ["overseas", "global", "hang seng", "china", "greater china", "world", "emerging market"], "priority": 7},
    {"tag": "Pharma/Healthcare", "contains": ["pharma", "healthcare", "health"], "priority": 8},
    {"tag": "Technology", "contains": ["tech", "artificial intelligence", "ai ", "fang", "digital"], "priority": 9},
    {"tag": "Commodities", "contains": ["commodit", "metal", "energy", "mining", "oil", "gas"], "priority": 10},
    {"tag": "Infrastructure", "contains": ["infra", "infrastructure"], "priority": 11},
    {"tag": "Banking/Financial", "contains": ["bank", "financial", "psu bank", "private bank"], "priority": 12},
    {"tag": "Index Fund", "contains": ["index", "nifty", "sensex", "bse", "midcap", "smallcap", "next 50", "nifty 50", "nifty50"], "priority": 13},
]

# ══════════════════════════════════════════════════════════════════════════════
# COLORS
# ══════════════════════════════════════════════════════════════════════════════
class Colors:
    TITLE_BG = "0D1117"
    TITLE_FG = "FFFFFF"
    INFO_BG = "F0F4F8"
    INFO_FG = "445566"
    BORDER = "D0D8E4"
    COL_HDR_BG = "1A3A5C"
    COL_HDR_FG = "FFFFFF"
    MOMENTUM_BG = "E8560A"
    MOMENTUM_FG = "FFFFFF"
    LONGTERM_BG = "1E6B8C"
    LONGTERM_FG = "FFFFFF"
    ENGINE_BG = "2D5016"
    ENGINE_FG = "FFFFFF"
    SIGNAL_BG = "6A1B9A"
    SIGNAL_FG = "FFFFFF"
    RANK1_BG = "FFD700"
    RANK2_BG = "E8E8E8"
    RANK3_BG = "D4956A"
    ALT_ROW = "F5F8FC"
    WHITE = "FFFFFF"
    POSITIVE = "1E7A4B"
    NEGATIVE = "C0392B"
    ENGINE1_TINT = "FFF3E0"
    ENGINE2_TINT = "E3F2FD"
    COMP_TINT = "F0FFF4"
    MOMENTUM_ONLY_BG = "FFF8E1"
    MOMENTUM_ONLY_FG = "E65100"
    MISSING_DATA_BG = "ECEFF1"
    MISSING_DATA_FG = "78909C"
    ASMP_TITLE = "0D1117"
    ASMP_BLEND = "2D5016"
    ASMP_ROW_BL = "EAFAF1"
    LEGEND_FULL = "E8F5E9"
    LEGEND_MOMENTUM = "FFF8E1"
    LEGEND_MISSING = "ECEFF1"
    CONSOLIDATED_BG = "1565C0"
    STATUS_FULL = "4CAF50"
    STATUS_MOMENTUM = "FF9800"
    STATUS_MISSING = "9E9E9E"
    SECTION_BLUE = "1565C0"
    SECTION_PURPLE = "7B1FA2"
    SECTION_TEAL = "00695C"
    TOP5_BG = "E3F2FD"
    ASSET_HDR_BG = "00695C"          # ★ FIX #2: asset-class group banner color

C = Colors()

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for col_key, value in FILTERS.items():
        actual = cols_lower.get(col_key.lower().strip())
        if actual:
            df = df[df[actual].astype(str).str.strip().str.lower() == str(value).strip().lower()]
    return df

def load_data() -> pd.DataFrame:
    sheets = pd.read_excel(CONFIG.INPUT_FILE, sheet_name=None)
    frames = [df for df in sheets.values() if "scheme_name" in df.columns]
    if not frames:
        raise ValueError("No valid sheets found with 'scheme_name' column")
    return apply_filters(pd.concat(frames, ignore_index=True))

# ══════════════════════════════════════════════════════════════════════════════
# SCORING ENGINE
# ══════════════════════════════════════════════════════════════════════════════
def to_num(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace('%', '', regex=False).str.replace(',', '', regex=False).str.strip()
    return pd.to_numeric(s, errors='coerce')

def pct_rank(series: pd.Series) -> pd.Series:
    valid = series.dropna()
    if valid.empty:
        return series.fillna(0.0)
    ranks = series.rank(method='min', na_option='bottom')
    return (ranks - 1) / max(len(series) - 1, 1) * 100

def cagr_2y(r: float) -> float:
    if pd.isna(r) or r <= -100:
        return 0.0
    return ((1 + float(r) / 100) ** 0.5 - 1) * 100

def assign_asset_tag(scheme_name: str) -> str:
    name_lower = scheme_name.lower()
    for rule in sorted(ASSET_TAG_RULES, key=lambda x: x.get("priority", 99)):
        for keyword in rule.get("contains", []):
            if keyword in name_lower:
                return rule["tag"]
    return "Standard Equity/Debt"

def calculate_trend_strength(row: pd.Series) -> Tuple[float, str]:
    r1m, r3m, r6m = row["_r1m"], row["_r3m"], row["_r6m"]
    if pd.isna(r1m) or pd.isna(r3m) or pd.isna(r6m):
        return 0, ""
    if r6m > r3m > r1m:
        return CONFIG.TREND_BONUS, "📈 Uptrend"
    elif r6m > r3m or r3m > r1m:
        return CONFIG.TREND_BONUS / 2, "↗️ Moderate"
    elif r6m < r3m < r1m:
        return -CONFIG.TREND_BONUS / 2, "📉 Downtrend"
    return 0, ""

def score_funds(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Returns are kept in PERCENT-POINT units internally for scoring/percentiles
    df["_r1m"] = to_num(df[COLUMN_MAP["return_1m"]]).fillna(0)
    df["_r3m"] = to_num(df[COLUMN_MAP["return_3m"]]).fillna(0)
    df["_r6m"] = to_num(df[COLUMN_MAP["return_6m"]]).fillna(0)
    df["_r1y"] = to_num(df[COLUMN_MAP["return_1y"]]).fillna(0)
    df["_r2y_raw"] = to_num(df[COLUMN_MAP["return_2y"]]).fillna(0)
    df["_r2y_cagr"] = df["_r2y_raw"].apply(cagr_2y)
    df["_r3y"] = to_num(df[COLUMN_MAP["return_3y"]]).fillna(0)
    df["_cat"] = df[COLUMN_MAP["category"]].astype(str).str.strip().str.title()
    df["_asset_class"] = df[COLUMN_MAP["scheme_name"]].apply(assign_asset_tag)

    df["_r1m_missing"] = to_num(df[COLUMN_MAP["return_1m"]]).isna()
    df["_r3m_missing"] = to_num(df[COLUMN_MAP["return_3m"]]).isna()
    df["_r6m_missing"] = to_num(df[COLUMN_MAP["return_6m"]]).isna()
    df["_r2y_missing"] = to_num(df[COLUMN_MAP["return_2y"]]).isna()
    df["_r3y_missing"] = to_num(df[COLUMN_MAP["return_3y"]]).isna()

    def classify_with_missing(row):
        has_momentum = not (row["_r1m_missing"] or row["_r3m_missing"] or row["_r6m_missing"])
        has_longterm = not (row["_r2y_missing"] or row["_r3y_missing"])
        if has_momentum and has_longterm:
            return "FULL"
        elif has_momentum and not has_longterm:
            return "MOMENTUM_ONLY"
        else:
            return "MISSING"

    df["_data_status"] = df.apply(classify_with_missing, axis=1)
    df["_has_missing_data"] = df["_data_status"] == "MISSING"

    mask_2y = df["_r2y_cagr"] > (QUALITY_FILTERS["cagr_2y_min"] * 100)
    mask_3y = df["_r3y"] > (QUALITY_FILTERS["cagr_3y_min"] * 100)
    mask_drawdown = df["_r1y"] > QUALITY_FILTERS["min_1y_return"]
    df["_qualifies"] = mask_2y & mask_3y & mask_drawdown & (df["_data_status"] == "FULL")

    trend_results = df.apply(calculate_trend_strength, axis=1)
    df["_trend_bonus"] = trend_results.apply(lambda x: x[0])
    df["_trend"] = trend_results.apply(lambda x: x[1])

    df["_e1"] = 0.0
    df["_e2"] = 0.0

    for cat in df["_cat"].unique():
        cm = df["_cat"] == cat
        cq = cm & df["_qualifies"]
        valid_m_mask = cm & (df["_data_status"].isin(["FULL", "MOMENTUM_ONLY"]))

        if valid_m_mask.sum() > 0:
            e1 = (pct_rank(df.loc[valid_m_mask, "_r6m"]) * ENGINE1_WEIGHTS["return_6m"] +
                  pct_rank(df.loc[valid_m_mask, "_r3m"]) * ENGINE1_WEIGHTS["return_3m"] +
                  pct_rank(df.loc[valid_m_mask, "_r1y"]) * ENGINE1_WEIGHTS["return_1y"] +
                  pct_rank(df.loc[valid_m_mask, "_r1m"]) * ENGINE1_WEIGHTS["return_1m"])
            e1 = e1 + df.loc[valid_m_mask, "_trend_bonus"]
            df.loc[valid_m_mask, "_e1"] = e1.clip(0, 100)

        if cq.sum() > 0:
            e2 = (pct_rank(df.loc[cq, "_r1y"]) * ENGINE2_WEIGHTS["return_1y"] +
                  pct_rank(df.loc[cq, "_r2y_cagr"]) * ENGINE2_WEIGHTS["return_2y"] +
                  pct_rank(df.loc[cq, "_r3y"]) * ENGINE2_WEIGHTS["return_3y"])
            df.loc[cq, "_e2"] = e2

    df["_comp"] = (df["_e1"] * COMPOSITE_BLEND["engine1_momentum"] +
                   df["_e2"] * COMPOSITE_BLEND["engine2_quality"])
    df.loc[df["_data_status"] == "MISSING", "_comp"] = 0.0
    df["_rank"] = (df.groupby("_cat")["_comp"]
                   .rank(method='min', ascending=False)
                   .astype(int))
    return df

# ══════════════════════════════════════════════════════════════════════════════
# SIGNAL FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def momentum_signal(e1: float, data_status: str = "FULL") -> str:
    if data_status == "MISSING":
        return "❌ N/A"
    if e1 >= 90:
        return "🔥 Hot"
    if e1 >= 75:
        return "⭐ Strong"
    if e1 >= 60:
        return "📈 Good"
    if e1 >= 40:
        return "➡️ Neutral"
    return "📉 Weak"

def quality_signal(e2: float, data_status: str = "FULL") -> str:
    if data_status == "MISSING":
        return "❌ N/A"
    if data_status == "MOMENTUM_ONLY":
        return "⏳ New Fund"
    if e2 >= 90:
        return "🏆 Elite"
    if e2 >= 75:
        return "⭐ Strong"
    if e2 >= 60:
        return "🏛️ Solid"
    if e2 >= 40:
        return "➡️ Average"
    if e2 > 0:
        return "⚠️ Below Avg"
    return "🔴 Not Qualified"

def composite_signal(comp: float, trend: str = "", e1: float = 0, e2: float = 0, data_status: str = "FULL") -> str:
    if data_status == "MISSING":
        return "❌ Missing Data"
    if data_status == "MOMENTUM_ONLY":
        if e1 >= 80:
            return "🔥 Hot Momentum"
        elif e1 >= 60:
            return "📈 Momentum Only"
        return "⏳ New Fund"
    if comp >= 85 and "Uptrend" in str(trend):
        return "🚀 Strong Conviction"
    if comp >= 75:
        return "⭐ Strong Buy"
    if comp >= 60:
        if e1 > e2:
            return "📈 Momentum Play"
        return "🏛️ Quality Hold"
    if comp >= 55:
        return "✅ Buy"
    if comp >= 40:
        return "⚠️ Watch"
    return "🔴 Avoid"

# ══════════════════════════════════════════════════════════════════════════════
# FORMATTING HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def border() -> Border:
    s = Side(style='thin', color=C.BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

# ★ FIX #1 — Convert a percent-point value (28.04) to a true fraction (0.2804)
#            so the genuine '0.00%' Excel format renders it correctly and the
#            stored number is mathematically consistent.
def pct_value(val, is_missing: bool = False) -> float:
    """Return a DECIMAL FRACTION for a true Excel percent format. 0 for missing."""
    if is_missing or pd.isna(val) or val is None:
        return 0.0
    try:
        return round(float(val) / 100.0, 6)   # 28.04 -> 0.2804  ★ FIX
    except Exception:
        return 0.0

def score_col(val) -> str:
    try:
        v = float(val)
        if v < 0:
            return C.MISSING_DATA_FG
    except Exception:
        return C.MISSING_DATA_FG
    if v >= 75:
        return C.POSITIVE
    if v >= 50:
        return "E67E00"
    return C.NEGATIVE

def clean_name(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', '', str(name))[:31]

def hfont(size: int = 9, color: str = "FFFFFF", bold: bool = True) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color)

def dfont(size: int = 9, bold: bool = False, color: str = "000000") -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color)

def get_row_style(row, row_num: int):
    data_status = row["_data_status"]
    rank = row["_rank"]
    if data_status == "MISSING":
        return (C.MISSING_DATA_BG, C.MISSING_DATA_FG, True, False)
    elif data_status == "MOMENTUM_ONLY":
        return (C.MOMENTUM_ONLY_BG, C.MOMENTUM_ONLY_FG, True, False)
    else:
        if rank == 1:
            return (C.RANK1_BG, "000000", False, True)
        elif rank == 2:
            return (C.RANK2_BG, "000000", False, True)
        elif rank == 3:
            return (C.RANK3_BG, "000000", False, True)
        else:
            bg = C.ALT_ROW if row_num % 2 == 0 else C.WHITE
            return (bg, "000000", False, False)

# ══════════════════════════════════════════════════════════════════════════════
# COLUMN LAYOUTS
# ══════════════════════════════════════════════════════════════════════════════
COL_HEADERS_CATEGORY = [
    "Rank", "Scheme Name", "AMC", "Asset Class",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Momentum)", "Engine 2\n(Quality)", "Composite\nScore",
    "Momentum\nSignal", "Quality\nSignal", "Composite\nSignal", "Data\nStatus"
]

COL_HEADERS_SUMMARY = [
    "Rank", "Asset Class", "Top Scheme", "AMC", "Category",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Momentum)", "Engine 2\n(Quality)", "Composite\nScore",
    "Momentum\nSignal", "Quality\nSignal", "Composite\nSignal"
]

COL_HEADERS_CONSOLIDATED = [
    "Rank", "Scheme Name", "AMC", "Category", "Asset Class",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Momentum)", "Engine 2\n(Quality)", "Composite\nScore",
    "Momentum\nSignal", "Quality\nSignal", "Composite\nSignal", "Data\nStatus"
]

# ★ FIX #2 — Asset Class Detail sheet uses the SAME schema as the Summary sheet
#            (Rank, Asset Class, Scheme, AMC, Category, returns, engines, signals)
COL_HEADERS_ASSET = [
    "Rank", "Asset Class", "Scheme Name", "AMC", "Category",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Momentum)", "Engine 2\n(Quality)", "Composite\nScore",
    "Momentum\nSignal", "Quality\nSignal", "Composite\nSignal"
]

COL_WIDTHS_CATEGORY = [6, 45, 20, 16, 9, 9, 9, 9, 9, 9, 12, 12, 12, 12, 12, 14, 12]
COL_WIDTHS_SUMMARY = [6, 16, 42, 20, 14, 9, 9, 9, 9, 9, 9, 12, 12, 12, 12, 12, 14]
COL_WIDTHS_CONSOLIDATED = [6, 42, 20, 14, 14, 9, 9, 9, 9, 9, 9, 12, 12, 12, 12, 12, 14, 12]
COL_WIDTHS_ASSET = [6, 18, 46, 22, 14, 9, 9, 9, 9, 9, 9, 12, 12, 12, 12, 12, 14]  # ★ FIX #2

RETURN_COLS_IDX_CAT = {5, 6, 7, 8, 9, 10}
RETURN_COLS_IDX_SUMMARY = {6, 7, 8, 9, 10, 11}
RETURN_COLS_IDX_CONSOLIDATED = {6, 7, 8, 9, 10, 11}
RETURN_COLS_IDX_ASSET = {6, 7, 8, 9, 10, 11}   # ★ FIX #2

MOMENTUM_COLS_CAT = (5, 7)
LONGTERM_COLS_CAT = (8, 10)
ENGINE_COLS_CAT = (11, 13)
SIGNAL_COLS_CAT = (14, 16)

# ══════════════════════════════════════════════════════════════════════════════
# BUILD CATEGORY SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_category_sheet(wb, cat, cat_df):
    ws = wb.create_sheet(clean_name(cat))
    bd = border()
    ncols = len(COL_HEADERS_CATEGORY)
    data_start_row = 5
    data_end_row = data_start_row + len(cat_df) - 1
    status_col = ncols

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"✦  {cat.upper()}  ✦"
    c.font = Font(name="Arial", bold=True, size=14, color=C.TITLE_FG)
    c.fill = fill(C.TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    e1_pct = int(COMPOSITE_BLEND["engine1_momentum"] * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_quality"] * 100)
    dynamic_formula = (
        f'="Dual Engine: {e1_pct}% Momentum + {e2_pct}% Quality | '
        f'✅ Full: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"FULL")&" | '
        f'⚠️ Momentum: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"MOMENTUM_ONLY")&" | '
        f'❌ Missing: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"MISSING")&" | '
        f'Total: "&COUNTA(B{data_start_row}:B{data_end_row})'
    )
    c = ws["A2"]
    c.value = dynamic_formula
    c.font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    c.fill = fill(C.INFO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    for ci in range(1, 5):
        ws.cell(row=3, column=ci).fill = fill(C.COL_HDR_BG)
        ws.cell(row=3, column=ci).border = bd

    for g_start, g_end, label, bg, fg in [
        (MOMENTUM_COLS_CAT[0], MOMENTUM_COLS_CAT[1], "◄  Momentum  ►", C.MOMENTUM_BG, C.MOMENTUM_FG),
        (LONGTERM_COLS_CAT[0], LONGTERM_COLS_CAT[1], "◄  Long-Term  ►", C.LONGTERM_BG, C.LONGTERM_FG),
        (ENGINE_COLS_CAT[0], ENGINE_COLS_CAT[1], "◄  Engine Scores  ►", C.ENGINE_BG, C.ENGINE_FG),
        (SIGNAL_COLS_CAT[0], SIGNAL_COLS_CAT[1], "◄  Signals  ►", C.SIGNAL_BG, C.SIGNAL_FG),
    ]:
        ws.merge_cells(f"{get_column_letter(g_start)}3:{get_column_letter(g_end)}3")
        cell = ws.cell(row=3, column=g_start, value=label)
        cell.font = hfont(color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(g_start, g_end + 1):
            ws.cell(row=3, column=ci).fill = fill(bg)
            ws.cell(row=3, column=ci).border = bd

    ws.cell(row=3, column=status_col).fill = fill(C.COL_HDR_BG)
    ws.cell(row=3, column=status_col).border = bd
    ws.row_dimensions[3].height = 18

    for ci, hdr in enumerate(COL_HEADERS_CATEGORY, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont()
        c.fill = fill(C.COL_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[4].height = 28

    for i, (_, row) in enumerate(cat_df.iterrows(), data_start_row):
        rank = row["_rank"]
        data_status = row["_data_status"]
        rbg, text_color, is_italic, is_bold = get_row_style(row, i)

        rank_val = rank if data_status != "MISSING" else 0
        e1_val = round(row["_e1"], 1)
        e2_val = round(row["_e2"], 1) if data_status == "FULL" else 0
        comp_val = round(row["_comp"], 1)

        mom_sig = momentum_signal(row["_e1"], data_status)
        qual_sig = quality_signal(row["_e2"], data_status)
        comp_sig = composite_signal(row["_comp"], row.get("_trend", ""), row["_e1"], row["_e2"], data_status)

        # ★ FIX #1: returns stored as fractions via pct_value()
        vals = [
            rank_val,
            row.get(COLUMN_MAP["scheme_name"], ""),
            row.get(COLUMN_MAP["amc"], ""),
            row["_asset_class"],
            pct_value(row["_r1m"]), pct_value(row["_r3m"]), pct_value(row["_r6m"]),
            pct_value(row["_r1y"]), pct_value(row["_r2y_cagr"]), pct_value(row["_r3y"]),
            e1_val, e2_val, comp_val,
            mom_sig, qual_sig, comp_sig,
            data_status
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border = bd
            c.fill = fill(rbg)
            c.alignment = Alignment(horizontal="left" if ci in {2, 3, 4} else "center", vertical="center")

            if ci in RETURN_COLS_IDX_CAT:
                c.number_format = PCT_FMT      # ★ FIX: true percent format
                if isinstance(val, (int, float)):
                    c.font = Font(name="Arial", bold=is_bold, size=9, color=C.POSITIVE if val >= 0 else C.NEGATIVE)
            elif ci == 11:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.ENGINE1_TINT)
                c.number_format = SCORE_FMT
            elif ci == 12:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.ENGINE2_TINT)
                c.number_format = SCORE_FMT
            elif ci == 13:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.COMP_TINT)
                c.number_format = SCORE_FMT
            elif ci in {14, 15, 16}:
                c.font = Font(name="Arial", bold=True, size=8, italic=is_italic)
            elif ci == 17:
                if data_status == "FULL":
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_FULL)
                    c.fill = fill(C.LEGEND_FULL)
                elif data_status == "MOMENTUM_ONLY":
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_MOMENTUM)
                    c.fill = fill(C.LEGEND_MOMENTUM)
                else:
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_MISSING)
                    c.fill = fill(C.LEGEND_MISSING)
            else:
                c.font = dfont(bold=is_bold)

        ws.row_dimensions[i].height = 16

    for ci, w in enumerate(COL_WIDTHS_CATEGORY, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{data_end_row}"
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# BUILD SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_summary(wb, df_scored):
    ws = wb.create_sheet("🏆 SUMMARY", 0)
    bd = border()
    ncols = len(COL_HEADERS_SUMMARY)
    asset_classes = sorted(df_scored["_asset_class"].unique())
    data_start_row = 5
    data_end_row = data_start_row + len(asset_classes) - 1

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = "MF INTELLIGENCE — ASSET CLASS RANKED SUMMARY"
    c.font = Font(name="Arial", bold=True, size=15, color=C.TITLE_FG)
    c.fill = fill(C.TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    e1_pct = int(COMPOSITE_BLEND["engine1_momentum"] * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_quality"] * 100)
    dynamic_formula = (
        f'="Top Performer per Asset Class | {e1_pct}% Momentum + {e2_pct}% Quality | '
        f'Total Asset Classes: "&COUNTA(B{data_start_row}:B{data_end_row})&" | '
        f'⭐ Strong Buy+: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"*Strong*")'
    )
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = dynamic_formula
    c.font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    c.fill = fill(C.INFO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    for ci in range(1, 6):
        ws.cell(row=3, column=ci).fill = fill(C.COL_HDR_BG)
        ws.cell(row=3, column=ci).border = bd

    for g_start, g_end, label, bg, fg in [
        (6, 8, "◄  Momentum  ►", C.MOMENTUM_BG, C.MOMENTUM_FG),
        (9, 11, "◄  Long-Term  ►", C.LONGTERM_BG, C.LONGTERM_FG),
        (12, 14, "◄  Engine Scores  ►", C.ENGINE_BG, C.ENGINE_FG),
        (15, 17, "◄  Signals  ►", C.SIGNAL_BG, C.SIGNAL_FG),
    ]:
        ws.merge_cells(f"{get_column_letter(g_start)}3:{get_column_letter(g_end)}3")
        cell = ws.cell(row=3, column=g_start, value=label)
        cell.font = hfont(color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(g_start, g_end + 1):
            ws.cell(row=3, column=ci).fill = fill(bg)
            ws.cell(row=3, column=ci).border = bd
    ws.row_dimensions[3].height = 18

    for ci, hdr in enumerate(COL_HEADERS_SUMMARY, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont()
        c.fill = fill(C.COL_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[4].height = 28

    rank_counter = 1
    for i, asset_class in enumerate(asset_classes, data_start_row):
        asset_df = df_scored[df_scored["_asset_class"] == asset_class].copy()
        asset_df = asset_df.sort_values(
            ["_data_status", "_comp"],
            ascending=[True, False],
            key=lambda x: x.map({"FULL": 0, "MOMENTUM_ONLY": 1, "MISSING": 2}) if x.name == "_data_status" else x
        )
        if asset_df.empty:
            continue

        top = asset_df.iloc[0]
        data_status = top["_data_status"]
        rbg, text_color, is_italic, is_bold = get_row_style(top, i)

        rank_val = rank_counter
        e1_val = round(top["_e1"], 1)
        e2_val = round(top["_e2"], 1) if data_status == "FULL" else 0
        comp_val = round(top["_comp"], 1)
        rank_counter += 1

        mom_sig = momentum_signal(top["_e1"], data_status)
        qual_sig = quality_signal(top["_e2"], data_status)
        comp_sig = composite_signal(top["_comp"], top.get("_trend", ""), top["_e1"], top["_e2"], data_status)

        # ★ FIX #1: returns stored as fractions
        vals = [
            rank_val,
            asset_class,
            top.get(COLUMN_MAP["scheme_name"], ""),
            top.get(COLUMN_MAP["amc"], ""),
            top["_cat"],
            pct_value(top["_r1m"]), pct_value(top["_r3m"]), pct_value(top["_r6m"]),
            pct_value(top["_r1y"]), pct_value(top["_r2y_cagr"]), pct_value(top["_r3y"]),
            e1_val, e2_val, comp_val,
            mom_sig, qual_sig, comp_sig
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border = bd
            c.fill = fill(rbg)
            c.alignment = Alignment(horizontal="left" if ci in {2, 3, 4, 5} else "center", vertical="center")

            if ci in RETURN_COLS_IDX_SUMMARY:
                c.number_format = PCT_FMT      # ★ FIX
                if isinstance(val, (int, float)):
                    c.font = Font(name="Arial", size=9, color=C.POSITIVE if val >= 0 else C.NEGATIVE)
            elif ci == 12:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.ENGINE1_TINT)
                c.number_format = SCORE_FMT
            elif ci == 13:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.ENGINE2_TINT)
                c.number_format = SCORE_FMT
            elif ci == 14:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.COMP_TINT)
                c.number_format = SCORE_FMT
            elif ci in {15, 16, 17}:
                c.font = Font(name="Arial", bold=True, size=8)
            else:
                c.font = dfont()

        ws.row_dimensions[i].height = 18

    for ci, w in enumerate(COL_WIDTHS_SUMMARY, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{data_end_row}"
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# ★ FIX #2 — NEW: BUILD ASSET CLASS DETAIL SHEET
# Standalone sheet matching the Summary look & feel, but showing the SAME
# detailed rows that exist in the individual asset-class sheets — every fund,
# grouped under an asset-class banner, ranked within its asset class.
# ══════════════════════════════════════════════════════════════════════════════
def build_asset_class_sheet(wb, df_scored):
    ws = wb.create_sheet("🎯 ASSET CLASS ANALYSIS", 2)   # placed right after Assumptions
    bd = border()
    ncols = len(COL_HEADERS_ASSET)

    # ---- Title (same dark banner as Summary) ----
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = "🎯 ASSET CLASS ANALYSIS — ALL FUNDS BY ASSET CLASS"
    c.font = Font(name="Arial", bold=True, size=15, color=C.TITLE_FG)
    c.fill = fill(C.TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ---- Dynamic info row (placeholder counts filled after we know last row) ----
    info_cell = ws["A2"]
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    info_cell.font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    info_cell.fill = fill(C.INFO_BG)
    info_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # ---- Group header band (row 3) identical to Summary ----
    for ci in range(1, 6):
        ws.cell(row=3, column=ci).fill = fill(C.COL_HDR_BG)
        ws.cell(row=3, column=ci).border = bd
    for g_start, g_end, label, bg, fg in [
        (6, 8, "◄  Momentum  ►", C.MOMENTUM_BG, C.MOMENTUM_FG),
        (9, 11, "◄  Long-Term  ►", C.LONGTERM_BG, C.LONGTERM_FG),
        (12, 14, "◄  Engine Scores  ►", C.ENGINE_BG, C.ENGINE_FG),
        (15, 17, "◄  Signals  ►", C.SIGNAL_BG, C.SIGNAL_FG),
    ]:
        ws.merge_cells(f"{get_column_letter(g_start)}3:{get_column_letter(g_end)}3")
        cell = ws.cell(row=3, column=g_start, value=label)
        cell.font = hfont(color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(g_start, g_end + 1):
            ws.cell(row=3, column=ci).fill = fill(bg)
            ws.cell(row=3, column=ci).border = bd
    ws.row_dimensions[3].height = 18

    # ---- Column headers (row 4) ----
    for ci, hdr in enumerate(COL_HEADERS_ASSET, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont()
        c.fill = fill(C.COL_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[4].height = 28

    # ---- Data: every fund, grouped & ranked within each asset class ----
    asset_classes = sorted(df_scored["_asset_class"].unique())
    r = 5
    data_start_row = 5

    for asset_class in asset_classes:
        asset_df = df_scored[df_scored["_asset_class"] == asset_class].copy()
        asset_df = asset_df.sort_values(
            ["_data_status", "_comp"],
            ascending=[True, False],
            key=lambda x: x.map({"FULL": 0, "MOMENTUM_ONLY": 1, "MISSING": 2}) if x.name == "_data_status" else x
        ).reset_index(drop=True)
        if asset_df.empty:
            continue

        # Asset-class banner row spanning all columns (teal, like the section bands)
        ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
        bc = ws.cell(row=r, column=1, value=f"►  {asset_class}   ({len(asset_df)} funds)")
        bc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        bc.fill = fill(C.ASSET_HDR_BG)
        bc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for ci in range(1, ncols + 1):
            ws.cell(row=r, column=ci).border = bd
            ws.cell(row=r, column=ci).fill = fill(C.ASSET_HDR_BG)
        ws.row_dimensions[r].height = 20
        r += 1

        for local_rank, (_, row) in enumerate(asset_df.iterrows(), 1):
            data_status = row["_data_status"]

            # rank style based on within-asset-class position
            if data_status == "MISSING":
                rbg, is_italic, is_bold = C.MISSING_DATA_BG, True, False
            elif data_status == "MOMENTUM_ONLY":
                rbg, is_italic, is_bold = C.MOMENTUM_ONLY_BG, True, False
            else:
                if local_rank == 1:
                    rbg, is_bold = C.RANK1_BG, True
                elif local_rank == 2:
                    rbg, is_bold = C.RANK2_BG, True
                elif local_rank == 3:
                    rbg, is_bold = C.RANK3_BG, True
                else:
                    rbg, is_bold = (C.ALT_ROW if r % 2 == 0 else C.WHITE), False
                is_italic = False

            e1_val = round(row["_e1"], 1)
            e2_val = round(row["_e2"], 1) if data_status == "FULL" else 0
            comp_val = round(row["_comp"], 1)

            mom_sig = momentum_signal(row["_e1"], data_status)
            qual_sig = quality_signal(row["_e2"], data_status)
            comp_sig = composite_signal(row["_comp"], row.get("_trend", ""), row["_e1"], row["_e2"], data_status)

            # ★ FIX #1: returns stored as fractions
            vals = [
                local_rank,
                asset_class,
                row.get(COLUMN_MAP["scheme_name"], ""),
                row.get(COLUMN_MAP["amc"], ""),
                row["_cat"],
                pct_value(row["_r1m"]), pct_value(row["_r3m"]), pct_value(row["_r6m"]),
                pct_value(row["_r1y"]), pct_value(row["_r2y_cagr"]), pct_value(row["_r3y"]),
                e1_val, e2_val, comp_val,
                mom_sig, qual_sig, comp_sig
            ]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.border = bd
                c.fill = fill(rbg)
                c.alignment = Alignment(horizontal="left" if ci in {2, 3, 4, 5} else "center", vertical="center")

                if ci in RETURN_COLS_IDX_ASSET:
                    c.number_format = PCT_FMT      # ★ FIX: true percent format
                    if isinstance(val, (int, float)):
                        c.font = Font(name="Arial", bold=is_bold, size=9, color=C.POSITIVE if val >= 0 else C.NEGATIVE)
                elif ci == 12:
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE1_TINT)
                    c.number_format = SCORE_FMT
                elif ci == 13:
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE2_TINT)
                    c.number_format = SCORE_FMT
                elif ci == 14:
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.COMP_TINT)
                    c.number_format = SCORE_FMT
                elif ci in {15, 16, 17}:
                    c.font = Font(name="Arial", bold=True, size=8, italic=is_italic)
                else:
                    c.font = dfont(bold=is_bold)

            ws.row_dimensions[r].height = 16
            r += 1

    data_end_row = r - 1

    # Fill the dynamic info row now that we know the data extent.
    # Composite Signal lives in column Q (17); we count Strong Buy+ across full range.
    info_cell.value = (
        f'="All Funds by Asset Class | {int(COMPOSITE_BLEND["engine1_momentum"]*100)}% Momentum + '
        f'{int(COMPOSITE_BLEND["engine2_quality"]*100)}% Quality | '
        f'Asset Classes: {len(asset_classes)} | '
        f'⭐ Strong Buy+: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"*Strong*")'
    )

    for ci, w in enumerate(COL_WIDTHS_ASSET, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# BUILD CONSOLIDATED SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_consolidated_sheet(wb, df_scored):
    ws = wb.create_sheet("📊 CONSOLIDATED")
    bd = border()
    ncols = len(COL_HEADERS_CONSOLIDATED)

    df_all = df_scored.copy()
    df_all = df_all.sort_values(
        ["_data_status", "_comp"],
        ascending=[True, False],
        key=lambda x: x.map({"FULL": 0, "MOMENTUM_ONLY": 1, "MISSING": 2}) if x.name == "_data_status" else x
    ).reset_index(drop=True)

    data_start_row = 5
    data_end_row = data_start_row + len(df_all) - 1
    status_col = ncols

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = "📊 CONSOLIDATED VIEW — ALL FUNDS RANKED"
    c.font = Font(name="Arial", bold=True, size=15, color=C.TITLE_FG)
    c.fill = fill(C.CONSOLIDATED_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    e1_pct = int(COMPOSITE_BLEND["engine1_momentum"] * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_quality"] * 100)
    dynamic_formula = (
        f'="All Funds Consolidated | {e1_pct}% Momentum + {e2_pct}% Quality | '
        f'✅ Full: "&COUNTIF(R{data_start_row}:R{data_end_row},"FULL")&" | '
        f'⚠️ Momentum: "&COUNTIF(R{data_start_row}:R{data_end_row},"MOMENTUM_ONLY")&" | '
        f'❌ Missing: "&COUNTIF(R{data_start_row}:R{data_end_row},"MISSING")&" | '
        f'Total: "&COUNTA(B{data_start_row}:B{data_end_row})'
    )
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = dynamic_formula
    c.font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    c.fill = fill(C.INFO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    for ci in range(1, 6):
        ws.cell(row=3, column=ci).fill = fill(C.COL_HDR_BG)
        ws.cell(row=3, column=ci).border = bd

    for g_start, g_end, label, bg, fg in [
        (6, 8, "◄  Momentum  ►", C.MOMENTUM_BG, C.MOMENTUM_FG),
        (9, 11, "◄  Long-Term  ►", C.LONGTERM_BG, C.LONGTERM_FG),
        (12, 14, "◄  Engine Scores  ►", C.ENGINE_BG, C.ENGINE_FG),
        (15, 17, "◄  Signals  ►", C.SIGNAL_BG, C.SIGNAL_FG),
    ]:
        ws.merge_cells(f"{get_column_letter(g_start)}3:{get_column_letter(g_end)}3")
        cell = ws.cell(row=3, column=g_start, value=label)
        cell.font = hfont(color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(g_start, g_end + 1):
            ws.cell(row=3, column=ci).fill = fill(bg)
            ws.cell(row=3, column=ci).border = bd

    ws.cell(row=3, column=status_col).fill = fill(C.COL_HDR_BG)
    ws.cell(row=3, column=status_col).border = bd
    ws.row_dimensions[3].height = 18

    for ci, hdr in enumerate(COL_HEADERS_CONSOLIDATED, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont()
        c.fill = fill(C.COL_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[4].height = 28

    global_rank = 1
    for i, (_, row) in enumerate(df_all.iterrows(), data_start_row):
        data_status = row["_data_status"]

        if data_status == "MISSING":
            rbg, is_italic, is_bold = C.MISSING_DATA_BG, True, False
        elif data_status == "MOMENTUM_ONLY":
            rbg, is_italic, is_bold = C.MOMENTUM_ONLY_BG, True, False
        else:
            if global_rank == 1:
                rbg, is_bold = C.RANK1_BG, True
            elif global_rank == 2:
                rbg, is_bold = C.RANK2_BG, True
            elif global_rank == 3:
                rbg, is_bold = C.RANK3_BG, True
            elif global_rank <= 10:
                rbg, is_bold = C.COMP_TINT, False
            else:
                rbg, is_bold = (C.ALT_ROW if i % 2 == 0 else C.WHITE), False
            is_italic = False

        rank_val = global_rank
        e1_val = round(row["_e1"], 1)
        e2_val = round(row["_e2"], 1) if data_status == "FULL" else 0
        comp_val = round(row["_comp"], 1)
        global_rank += 1

        mom_sig = momentum_signal(row["_e1"], data_status)
        qual_sig = quality_signal(row["_e2"], data_status)
        comp_sig = composite_signal(row["_comp"], row.get("_trend", ""), row["_e1"], row["_e2"], data_status)

        # ★ FIX #1: returns stored as fractions
        vals = [
            rank_val,
            row.get(COLUMN_MAP["scheme_name"], ""),
            row.get(COLUMN_MAP["amc"], ""),
            row["_cat"],
            row["_asset_class"],
            pct_value(row["_r1m"]), pct_value(row["_r3m"]), pct_value(row["_r6m"]),
            pct_value(row["_r1y"]), pct_value(row["_r2y_cagr"]), pct_value(row["_r3y"]),
            e1_val, e2_val, comp_val,
            mom_sig, qual_sig, comp_sig,
            data_status
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border = bd
            c.fill = fill(rbg)
            c.alignment = Alignment(horizontal="left" if ci in {2, 3, 4, 5} else "center", vertical="center")

            if ci in RETURN_COLS_IDX_CONSOLIDATED:
                c.number_format = PCT_FMT      # ★ FIX
                if isinstance(val, (int, float)):
                    c.font = Font(name="Arial", bold=is_bold, size=9, color=C.POSITIVE if val >= 0 else C.NEGATIVE)
            elif ci == 12:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.ENGINE1_TINT)
                c.number_format = SCORE_FMT
            elif ci == 13:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.ENGINE2_TINT)
                c.number_format = SCORE_FMT
            elif ci == 14:
                c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                c.fill = fill(C.COMP_TINT)
                c.number_format = SCORE_FMT
            elif ci in {15, 16, 17}:
                c.font = Font(name="Arial", bold=True, size=8, italic=is_italic)
            elif ci == 18:
                if data_status == "FULL":
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_FULL)
                    c.fill = fill(C.LEGEND_FULL)
                elif data_status == "MOMENTUM_ONLY":
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_MOMENTUM)
                    c.fill = fill(C.LEGEND_MOMENTUM)
                else:
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_MISSING)
                    c.fill = fill(C.LEGEND_MISSING)
            else:
                c.font = dfont(bold=is_bold)

        ws.row_dimensions[i].height = 16

    for ci, w in enumerate(COL_WIDTHS_CONSOLIDATED, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{data_end_row}"
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# BUILD COMPREHENSIVE ASSUMPTIONS SHEET
# (Section 12 "Top 5 per Asset Class" REMOVED — that content now lives in the
#  dedicated 🎯 ASSET CLASS ANALYSIS sheet.)   ★ FIX #2
# ══════════════════════════════════════════════════════════════════════════════
def build_assumptions(wb, df_scored):
    ws = wb.create_sheet("📋 ASSUMPTIONS", 1)
    bd = border()

    def section(row, title, bg, fg="FFFFFF", ncols=4):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws[f"A{row}"]
        c.value = title
        c.font = Font(name="Arial", bold=True, size=11, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        return row + 1

    def kv_row(row, key, val, desc="", extra="", row_bg="FFFFFF", ncols=4):
        values = [(1, key), (2, val), (3, desc)]
        if ncols >= 4:
            values.append((4, extra))
        for ci, text in values:
            c = ws.cell(row=row, column=ci, value=text)
            c.font = Font(name="Arial", size=9)
            c.fill = fill(row_bg)
            c.border = bd
            c.alignment = Alignment(horizontal="left" if ci in {1, 3, 4} else "center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18
        return row + 1

    def col_headers(row, headers, bg):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c.fill = fill(bg)
            c.border = bd
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 20
        return row + 1

    ws.merge_cells("A1:D1")
    ws["A1"] = "DUAL ENGINE MODEL — COMPLETE METHODOLOGY & ASSUMPTIONS"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill(C.ASMP_TITLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    # ★ FIX #1: note updated — values now stored as true decimal fractions
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Comprehensive Documentation | Returns stored as decimal fractions w/ true % format (0 = Missing)"
    ws["A2"].font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    ws["A2"].fill = fill(C.INFO_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    r = 4

    r = section(r, "🔍  DATA FILTERS (Pre-Processing)", C.SECTION_BLUE)
    r = col_headers(r, ["Filter Parameter", "Value", "Description", "Python Logic"], C.SECTION_BLUE)
    r = kv_row(r, "cat_level_1", "Open Ended Schemes", "Scheme category level 1", "df[col] == value", C.ASMP_ROW_BL)
    r = kv_row(r, "cat_level_2", "Other Scheme", "Scheme category level 2", "df[col] == value", C.ASMP_ROW_BL)
    r = kv_row(r, "plan_type", "Regular", "Regular plans only (not Direct)", "df[col] == value", C.ASMP_ROW_BL)
    r = kv_row(r, "option_type", "Growth", "Growth option only", "df[col] == value", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "📋  COLUMN MAPPING (Source → Internal)", C.SECTION_BLUE)
    r = col_headers(r, ["Internal Name", "Source Column", "Data Type", "Notes"], C.SECTION_BLUE)
    r = kv_row(r, "scheme_name", "scheme_name", "String", "Fund name for display", C.ASMP_ROW_BL)
    r = kv_row(r, "category", "cat_level_3", "String", "Category for grouping", C.ASMP_ROW_BL)
    r = kv_row(r, "amc", "amc_name", "String", "Asset Management Company", C.ASMP_ROW_BL)
    r = kv_row(r, "return_1m", "return_30d", "Numeric %", "30-day return", C.ASMP_ROW_BL)
    r = kv_row(r, "return_3m", "return_90d", "Numeric %", "90-day return", C.ASMP_ROW_BL)
    r = kv_row(r, "return_6m", "return_180d", "Numeric %", "180-day return", C.ASMP_ROW_BL)
    r = kv_row(r, "return_1y", "return_365d", "Numeric %", "365-day return", C.ASMP_ROW_BL)
    r = kv_row(r, "return_2y", "return_730d", "Numeric %", "730-day cumulative → CAGR", C.ASMP_ROW_BL)
    r = kv_row(r, "return_3y", "return_1095d", "Numeric %", "1095-day CAGR", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "📊  DATA STATUS CLASSIFICATION", C.ASMP_BLEND)
    r = col_headers(r, ["Status", "Condition", "Visual Style", "Scoring Impact"], C.ASMP_BLEND)
    r = kv_row(r, "✅ FULL", "All 1M/3M/6M + 2Y/3Y present", "White/Ranked Colors", "Both E1 and E2 calculated", C.LEGEND_FULL)
    r = kv_row(r, "⚠️ MOMENTUM_ONLY", "1M/3M/6M present, 2Y/3Y missing", "Light Amber + Italic", "Only E1 calculated, E2=0", C.LEGEND_MOMENTUM)
    r = kv_row(r, "❌ MISSING", "Missing any of 1M/3M/6M", "Light Gray + Italic", "E1=0, E2=0, Comp=0", C.LEGEND_MISSING)
    r += 1

    r = section(r, "🚦  QUALITY GATES (Engine 2 Eligibility)", C.SECTION_PURPLE)
    r = col_headers(r, ["Gate", "Threshold", "Formula", "Purpose"], C.SECTION_PURPLE)
    r = kv_row(r, "Min 2Y CAGR", f"{int(QUALITY_FILTERS['cagr_2y_min']*100)}%", "r2y_cagr > 10%", "Filter low long-term performers", C.ASMP_ROW_BL)
    r = kv_row(r, "Min 3Y CAGR", f"{int(QUALITY_FILTERS['cagr_3y_min']*100)}%", "r3y > 12%", "Filter inconsistent performers", C.ASMP_ROW_BL)
    r = kv_row(r, "Drawdown Tolerance", f"{int(QUALITY_FILTERS['min_1y_return'])}%", "r1y > -30%", "Exclude extreme losers", C.ASMP_ROW_BL)
    r = kv_row(r, "Data Status", "FULL only", "data_status == 'FULL'", "Only complete data qualifies", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "📈  ENGINE 1: MOMENTUM SCORE (Short-Term)", C.MOMENTUM_BG)
    r = col_headers(r, ["Component", "Weight", "Formula", "Description"], C.MOMENTUM_BG)
    r = kv_row(r, "6M Return Rank", f"{int(ENGINE1_WEIGHTS['return_6m']*100)}%", "pct_rank(r6m) × 0.30", "6-month percentile within category", C.ASMP_ROW_BL)
    r = kv_row(r, "3M Return Rank", f"{int(ENGINE1_WEIGHTS['return_3m']*100)}%", "pct_rank(r3m) × 0.20", "3-month percentile within category", C.ASMP_ROW_BL)
    r = kv_row(r, "1Y Return Rank", f"{int(ENGINE1_WEIGHTS['return_1y']*100)}%", "pct_rank(r1y) × 0.25", "1-year percentile within category", C.ASMP_ROW_BL)
    r = kv_row(r, "1M Return Rank", f"{int(ENGINE1_WEIGHTS['return_1m']*100)}%", "pct_rank(r1m) × 0.25", "1-month percentile within category", C.ASMP_ROW_BL)
    r = kv_row(r, "Trend Bonus", f"+{CONFIG.TREND_BONUS}", "If r6m > r3m > r1m", "📈 Uptrend bonus (0 to 5 pts)", C.ASMP_ROW_BL)
    r = kv_row(r, "Final E1", "0-100", "clip(sum + trend, 0, 100)", "Capped percentile score", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "🏛️  ENGINE 2: QUALITY SCORE (Long-Term)", C.LONGTERM_BG)
    r = col_headers(r, ["Component", "Weight", "Formula", "Description"], C.LONGTERM_BG)
    r = kv_row(r, "1Y Return Rank", f"{int(ENGINE2_WEIGHTS['return_1y']*100)}%", "pct_rank(r1y) × 0.25", "1-year percentile (qualified only)", C.ASMP_ROW_BL)
    r = kv_row(r, "2Y CAGR Rank", f"{int(ENGINE2_WEIGHTS['return_2y']*100)}%", "pct_rank(r2y_cagr) × 0.30", "2-year CAGR percentile", C.ASMP_ROW_BL)
    r = kv_row(r, "3Y CAGR Rank", f"{int(ENGINE2_WEIGHTS['return_3y']*100)}%", "pct_rank(r3y) × 0.45", "3-year CAGR percentile (highest weight)", C.ASMP_ROW_BL)
    r = kv_row(r, "Final E2", "0-100", "sum of weighted ranks", "Only for qualified funds", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "🎯  COMPOSITE SCORE CALCULATION", C.ENGINE_BG)
    r = col_headers(r, ["Component", "Weight", "Formula", "Description"], C.ENGINE_BG)
    r = kv_row(r, "Engine 1 Contribution", f"{int(COMPOSITE_BLEND['engine1_momentum']*100)}%", "E1 × 0.55", "Momentum component", C.ASMP_ROW_BL)
    r = kv_row(r, "Engine 2 Contribution", f"{int(COMPOSITE_BLEND['engine2_quality']*100)}%", "E2 × 0.45", "Quality component", C.ASMP_ROW_BL)
    r = kv_row(r, "Final Composite", "0-100", "E1×0.55 + E2×0.45", "Blended score for ranking", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "🏆  RANKING METHODOLOGY", C.SECTION_TEAL)
    r = col_headers(r, ["Step", "Logic", "Scope", "Result"], C.SECTION_TEAL)
    r = kv_row(r, "1. Group by Category", "df.groupby('_cat')", "Within each category", "Separate ranking pools", C.ASMP_ROW_BL)
    r = kv_row(r, "2. Sort by Composite", "descending order", "Highest score = Rank 1", "FULL data prioritized", C.ASMP_ROW_BL)
    r = kv_row(r, "3. Handle Ties", "method='min'", "Same score = same rank", "Preserves order stability", C.ASMP_ROW_BL)
    r = kv_row(r, "4. Data Status Sort", "FULL → MOMENTUM → MISSING", "Global consolidated view", "Best data quality first", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "🏷️  ASSET CLASS TAGGING (Priority Order)", C.SECTION_BLUE)
    r = col_headers(r, ["Priority", "Asset Class", "Keywords", "Example Match"], C.SECTION_BLUE)
    for rule in ASSET_TAG_RULES:
        keywords = ", ".join(rule["contains"][:3]) + ("..." if len(rule["contains"]) > 3 else "")
        r = kv_row(r, str(rule["priority"]), rule["tag"], keywords, "scheme_name.lower().contains()", C.ASMP_ROW_BL)
    r = kv_row(r, "99", "Standard Equity/Debt", "(default)", "No keyword match", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "📈  MOMENTUM SIGNAL (Engine 1)", C.MOMENTUM_BG)
    r = col_headers(r, ["Signal", "E1 Range", "Meaning", "Action"], C.MOMENTUM_BG)
    r = kv_row(r, "🔥 Hot", "90-100", "Exceptional momentum", "Strong short-term performer", C.ASMP_ROW_BL)
    r = kv_row(r, "⭐ Strong", "75-89", "Above average momentum", "Good entry point", C.ASMP_ROW_BL)
    r = kv_row(r, "📈 Good", "60-74", "Decent momentum", "Monitor for improvement", C.ASMP_ROW_BL)
    r = kv_row(r, "➡️ Neutral", "40-59", "Average performance", "No clear direction", C.ASMP_ROW_BL)
    r = kv_row(r, "📉 Weak", "<40", "Below average", "Caution advised", C.ASMP_ROW_BL)
    r += 1

    r = section(r, "🏛️  QUALITY SIGNAL (Engine 2)", C.LONGTERM_BG)
    r = col_headers(r, ["Signal", "E2 Range", "Meaning", "Action"], C.LONGTERM_BG)
    r = kv_row(r, "🏆 Elite", "90-100", "Top-tier quality", "Core holding candidate", C.ASMP_ROW_BL)
    r = kv_row(r, "⭐ Strong", "75-89", "Strong fundamentals", "Solid long-term choice", C.ASMP_ROW_BL)
    r = kv_row(r, "🏛️ Solid", "60-74", "Good quality metrics", "Reliable performer", C.ASMP_ROW_BL)
    r = kv_row(r, "➡️ Average", "40-59", "Market average", "No standout quality", C.ASMP_ROW_BL)
    r = kv_row(r, "⚠️ Below Avg", "1-39", "Below market average", "Review before investing", C.ASMP_ROW_BL)
    r = kv_row(r, "🔴 Not Qualified", "0", "Failed quality gates", "Does not meet criteria", C.ASMP_ROW_BL)
    r = kv_row(r, "⏳ New Fund", "N/A", "Insufficient history", "Wait for track record", C.LEGEND_MOMENTUM)
    r += 1

    r = section(r, "🎯  COMPOSITE SIGNAL (Combined)", C.ENGINE_BG)
    r = col_headers(r, ["Signal", "Condition", "Meaning", "Action"], C.ENGINE_BG)
    r = kv_row(r, "🚀 Strong Conviction", "≥85 + Uptrend", "Best of both engines", "Top pick", C.ASMP_ROW_BL)
    r = kv_row(r, "⭐ Strong Buy", "75-84", "High composite score", "Recommended buy", C.ASMP_ROW_BL)
    r = kv_row(r, "📈 Momentum Play", "60-74 (E1>E2)", "Momentum-driven", "Short-term opportunity", C.ASMP_ROW_BL)
    r = kv_row(r, "🏛️ Quality Hold", "60-74 (E2>E1)", "Quality-driven", "Long-term hold", C.ASMP_ROW_BL)
    r = kv_row(r, "✅ Buy", "55-59", "Moderate confidence", "Consider buying", C.ASMP_ROW_BL)
    r = kv_row(r, "⚠️ Watch", "40-54", "Below average", "Monitor closely", C.ASMP_ROW_BL)
    r = kv_row(r, "🔴 Avoid", "<40", "Poor performance", "Not recommended", C.ASMP_ROW_BL)
    r = kv_row(r, "🔥 Hot Momentum", "E1≥80 (Mom. Only)", "Strong momentum, no history", "Speculative", C.LEGEND_MOMENTUM)
    r += 1

    r = section(r, "📈  DATA STATISTICS (Live)", C.ASMP_BLEND)
    r = col_headers(r, ["Metric", "Count", "Formula", "Notes"], C.ASMP_BLEND)

    total_formula = '=COUNTA(\'📊 CONSOLIDATED\'!B5:B10000)'
    full_formula = '=COUNTIF(\'📊 CONSOLIDATED\'!R5:R10000,"FULL")'
    momentum_formula = '=COUNTIF(\'📊 CONSOLIDATED\'!R5:R10000,"MOMENTUM_ONLY")'
    missing_formula = '=COUNTIF(\'📊 CONSOLIDATED\'!R5:R10000,"MISSING")'

    for key, formula, note, bg_color in [
        ("Total Funds", total_formula, "All funds in analysis", C.ASMP_ROW_BL),
        ("✅ Full Data", full_formula, "Complete data for scoring", C.LEGEND_FULL),
        ("⚠️ Momentum Only", momentum_formula, "New funds, E2=0", C.LEGEND_MOMENTUM),
        ("❌ Missing Data", missing_formula, "Incomplete, Comp=0", C.LEGEND_MISSING),
    ]:
        c1 = ws.cell(row=r, column=1, value=key)
        c1.font = Font(name="Arial", size=9)
        c1.fill = fill(bg_color)
        c1.border = bd

        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = Font(name="Arial", size=9, bold=True)
        c2.fill = fill(bg_color)
        c2.border = bd
        c2.alignment = Alignment(horizontal="center")

        # ★ FIX: percentage share via a true percent format instead of TEXT()
        if key != "Total Funds":
            c3 = ws.cell(row=r, column=3, value=f'={formula}/{total_formula}')
            c3.number_format = '0.0%'
        else:
            c3 = ws.cell(row=r, column=3, value=1)
            c3.number_format = '0%'
        c3.font = Font(name="Arial", size=9)
        c3.fill = fill(bg_color)
        c3.border = bd
        c3.alignment = Alignment(horizontal="center")

        c4 = ws.cell(row=r, column=4, value=note)
        c4.font = Font(name="Arial", size=9)
        c4.fill = fill(bg_color)
        c4.border = bd

        ws.row_dimensions[r].height = 18
        r += 1

    r += 1

    # ★ FIX #2: Section 12 (Top 5 per Asset Class) removed.
    #           Pointer note added directing users to the new dedicated sheet.
    r = section(r, "🎯  ASSET CLASS ANALYSIS", C.CONSOLIDATED_BG)
    r = kv_row(r, "See dedicated sheet",
               "🎯 ASSET CLASS ANALYSIS",
               "All funds grouped & ranked by asset class (Summary look & feel)",
               "Detailed rows like the individual asset-class sheets", C.ASMP_ROW_BL)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 40

# ══════════════════════════════════════════════════════════════════════════════
# MAIN EXECUTOR
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("🚀 Loading data matrix...")
    df = load_data()

    print("⚙️ Executing Scoring Engine...")
    df_scored = score_funds(df)

    print("\n📊 Data Status Summary:")
    print(f"   ✅ Full Data: {len(df_scored[df_scored['_data_status'] == 'FULL'])}")
    print(f"   ⚠️ Momentum Only: {len(df_scored[df_scored['_data_status'] == 'MOMENTUM_ONLY'])}")
    print(f"   ❌ Missing Data: {len(df_scored[df_scored['_data_status'] == 'MISSING'])}")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    categories = sorted(df_scored["_cat"].unique())

    print("\n📊 Building Summary Sheet...")
    build_summary(wb, df_scored)

    print("📋 Building Assumptions Sheet...")
    build_assumptions(wb, df_scored)

    print("🎯 Building Asset Class Analysis Sheet...")   # ★ FIX #2
    build_asset_class_sheet(wb, df_scored)

    print("📁 Generating Category Sheets...")
    for cat in categories:
        cat_df = df_scored[df_scored["_cat"] == cat].sort_values("_rank")
        if not cat_df.empty:
            build_category_sheet(wb, cat, cat_df)

    print("📊 Building Consolidated Sheet...")
    build_consolidated_sheet(wb, df_scored)

    wb.save(CONFIG.OUTPUT_FILE)
    print(f"\n✅ Done! Output: '{CONFIG.OUTPUT_FILE}'")
    print("   📊 Summary: Asset Class rankings")
    print("   📋 Assumptions: Complete methodology")
    print("   🎯 Asset Class Analysis: All funds by asset class (NEW)")
    print("   📁 Category Sheets: True percent values")
    print(f"   📊 CONSOLIDATED: All {len(df_scored)} funds")

if __name__ == "__main__":
    main()
